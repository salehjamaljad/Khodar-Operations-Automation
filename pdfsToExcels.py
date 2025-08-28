import os
import pandas as pd
import pdfplumber
import zipfile
import tempfile
from io import BytesIO
from fuzzywuzzy import process
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import re

def process_talabat_invoices(
    zip_file_bytes: bytes,
    invoice_date: str,
    base_invoice_number: int,
    translation_dict: dict,
    categories_dict: dict,
    branches_dict: dict,
    branches_translation_tlbt: dict,
    columns: list
) -> bytes:
    standardized_columns = [col.replace("\n", "_") for col in columns]
    selected_date = invoice_date  # string in "YYYY-MM-DD"

    special_codes = {
        "EG_Alex East_DS_", "EG_Alex", "EG_Zahraa Maadi", "EG_Nasrcity", "EG_Mansoura",
        "EG_Tagamoa Golden", "EG_Tagamoa", "EG_Madinaty", "EG_Hadayek", "EG_October", "EG_Shrouk_", "EG_Mokatam", "EG_Sheikh", "EG_Faisal"
    }

    def extract_eg_codes(pdf_path):
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + " "
            words = text.split()
            i = 0
            results = []
            while i < len(words):
                word = words[i]
                if word.startswith("EG_"):
                    if any(word == code or word.startswith(code) for code in special_codes):
                        next_word = words[i + 1] if i + 1 < len(words) else ""
                        combined = f"{word} {next_word}"
                        closest_match, score = process.extractOne(combined, branches_dict.keys())
                        if score >= 80:
                            results.append({
                                "filename": os.path.basename(pdf_path),
                                "extracted": combined,
                                "matched_key": closest_match,
                                "arabic_name": branches_dict[closest_match]
                            })
                        else:
                            results.append({"filename": os.path.basename(pdf_path), "extracted": combined})
                        i += 1
                    else:
                        closest_match, score = process.extractOne(word, branches_dict.keys())
                        if score >= 80:
                            results.append({
                                "filename": os.path.basename(pdf_path),
                                "extracted": word,
                                "matched_key": closest_match,
                                "arabic_name": branches_dict[closest_match]
                            })
                        else:
                            results.append({"filename": os.path.basename(pdf_path), "extracted": word})
                i += 1
            return results

    def process_pdf(file_path):
        all_tables = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    df = pd.DataFrame(table)
                    all_tables.append(df)

        for i, table in enumerate(all_tables):
            non_null_counts = table.notnull().sum()
            threshold = non_null_counts.max() * 0.5
            columns_to_drop = non_null_counts[non_null_counts <= threshold].index
            if len(columns_to_drop) > 0:
                all_tables[i] = table.drop(columns=columns_to_drop)

        for i, df in enumerate(all_tables[1:]):
            df.columns = standardized_columns

        if len(all_tables) > 1:
            final_df = pd.concat(all_tables[1:], ignore_index=True)
        else:
            final_df = all_tables[0]

        df = final_df
        df = df.loc[~(df.applymap(lambda x: x == "").all(axis=1))]
        df = df.reset_index(drop=True)
        df = df[df["Qty"] != ""]
        df = df[df["SKU"] != "SKU"]
        df.drop(
            columns=[
                "Disc._Amt.",
                "Amt._Excl._VAT",
                "VAT_%",
                "VAT_Amt.",
                "Supplier SKU",
                "No.",
                "Product",
            ],
            inplace=True,
        )

        df.rename(columns={"Unit_Cost": "PP", "Amt._Incl._VAT": "Total"}, inplace=True)

        df["PP"] = df["PP"].astype(float)
        df["Total"] = df["Total"].astype(float)
        df["Qty"] = df["Qty"].astype(int)
        try:
            df["Barcode"] = df["Barcode"].astype(int)
        except OverflowError:
            df["Barcode"] = df["Barcode"].astype(float)
        df["SKU"] = df["SKU"].astype(int)
        df["Item Name Ar"] = df["SKU"].map(translation_dict)
        df = df[["SKU", "Barcode", "Item Name Ar", "PP", "Qty", "Total"]]
        df = df.reset_index(drop=True)
        return df

    with tempfile.TemporaryDirectory() as temp_dir:
        zip_path = os.path.join(temp_dir, "uploaded.zip")
        with open(zip_path, "wb") as f:
            f.write(zip_file_bytes)

        with zipfile.ZipFile(zip_path, "r") as zip_ref:
            zip_ref.extractall(temp_dir)

        output_dir = os.path.join(temp_dir, "excels")
        os.makedirs(output_dir, exist_ok=True)

        # Step 1: Process each PDF → save an Excel + create "فاتورة" sheet
        pos_with_filenames = {}
        for filename in os.listdir(temp_dir):
            if not filename.endswith(".pdf"):
                continue
            file_path = os.path.join(temp_dir, filename)
            df = process_pdf(file_path)

            match = re.search(r"(PO\d+)", filename)
            po = match.group(1) if match else None
            pos_with_filenames[filename] = po

            extracted_data = extract_eg_codes(file_path)
            branch_name = None
            if extracted_data:
                branch_name = extracted_data[0].get("arabic_name", None)

            if branch_name:
                output_filename = f"{branch_name}_{po}_{selected_date}.xlsx"
            else:
                output_filename = f"{os.path.splitext(filename)[0]}.xlsx"

            output_path = os.path.join(output_dir, output_filename)
            df.to_excel(output_path, index=False, engine="openpyxl")

            wb = load_workbook(output_path)
            ws = wb.active
            ws["H1"] = po

            # Clear Qty & Total columns in df
            for col in df.columns:
                if col.strip().lower() == "qty":
                    df[col] = ""
                if col.strip().lower() == "total":
                    df[col] = ""

            # Create (or recreate) "فاتورة" sheet
            if "فاتورة" in wb.sheetnames:
                del wb["فاتورة"]
            ws_invoice = wb.create_sheet("فاتورة")

            thick_border = Border(
                left=Side(style="thick"),
                right=Side(style="thick"),
                top=Side(style="thick"),
                bottom=Side(style="thick"),
            )
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Write DataFrame rows starting at row 11
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=11):
                ws_invoice.row_dimensions[r_idx].height = 21
                for c_idx, value in enumerate(row, start=1):
                    cell = ws_invoice.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if df.columns[c_idx - 1].lower() == "barcode" and isinstance(value, (int, float)):
                        cell.number_format = "0"

            # Static cells + image
            img = Image("Picture1.png")
            ws_invoice.add_image(img, "A1")
            ws_invoice["F1"] = "فاتورة مبيعات"
            ws_invoice["F2"] = "رقم الفاتورة #"
            ws_invoice["F3"] = "تاريخ الاستلام "
            ws_invoice["E3"] = selected_date
            ws_invoice["F4"] = "امر شراء رقم"
            ws_invoice["E4"] = po
            ws_invoice["F6"] = "اسم العميل "
            ws_invoice["E6"] = "دليفيري هيرو ديمارت ايجيبت"
            ws_invoice["F7"] = "الفرع"
            ws_invoice["E7"] = branch_name
            ws_invoice["C1"] = "شركه خضار للتجارة والتسويق"
            ws_invoice["C1"].alignment = Alignment(horizontal="center", vertical="center")
            ws_invoice["C2"] = "Khodar for Trading & Marketing"
            ws_invoice["C2"].alignment = Alignment(horizontal="center", vertical="center")
            ws_invoice["A5"] = "خضار.كوم"

            for cell_ref in [
                "F1",
                "F2",
                "F3",
                "E3",
                "F4",
                "E4",
                "F6",
                "E6",
                "F7",
                "E7",
                "E2",
                "C1",
                "C2",
                "A5",
            ]:
                ws_invoice[cell_ref].border = thick_border

            # Adjust column widths and bold formatting
            df_end_row = 11 + len(df) + 1
            for col in ws_invoice.columns:
                max_length = 0
                column = col[0].column
                column_letter = get_column_letter(column)
                if column_letter == "A":
                    ws_invoice.column_dimensions[column_letter].width = 10
                else:
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws_invoice.column_dimensions[column_letter].width = max_length + 2
            for row in ws_invoice.iter_rows():
                for cell in row:
                    if cell.value or cell.coordinate == "E2":
                        cell.font = Font(bold=True)

            wb.save(output_path)

        # Step 2: Build combined DataFrame from all generated Excel files
        all_dfs = []
        for excel_file in os.listdir(output_dir):
            if excel_file.endswith(".xlsx"):
                excel_path = os.path.join(output_dir, excel_file)
                df = pd.read_excel(excel_path, usecols=range(6))
                base = os.path.splitext(excel_file)[0]
                parts = base.split("_")
                if len(parts) >= 2:
                    branch_name = parts[0]
                    po = parts[1]
                    df["branch"] = branch_name
                    df["po"] = po
                all_dfs.append(df)

        if all_dfs:
            combined_df = pd.concat(all_dfs, ignore_index=True)
            combined_df["SKU"] = pd.to_numeric(combined_df["SKU"], errors="coerce").astype("Int64")
            combined_df["Product"] = combined_df["SKU"].map(translation_dict)
            reverse_categories = {
                item: category for category, items in categories_dict.items() for item in items
            }
            combined_df["category"] = combined_df["Product"].map(reverse_categories)

            pivot_df = combined_df.pivot_table(
                index=["Barcode", "SKU", "Product", "category", "PP"],
                columns="branch",
                values="Qty",
                aggfunc="sum",
                fill_value=0,
            ).reset_index()
            pivot_df = pivot_df.rename(columns={"Product": "Product name"})

            alexandria_columns = [
                "Barcode",
                "Product name",
                "SKU",
                "category",
                "PP",
                "سيدي بشر",
                "الابراهيميه",
                "وينجت",
            ]
            ready_veg_columns = [
                "Barcode",
                "Product name",
                "SKU",
                "category",
                "PP",
                "المعادي لاسلكي",
                "الدقي",
                "زهراء المعادي",
                "ميدان لبنان",
                "العجوزة",
                "كورنيش المعادي",
                "زهراء المعادي - 2",
                "الظاهر",
                "المقطم",
                "السيدة زينب",
                "حلوان",
                "المنيل",
                "المقطم 2 هضبة",
                "شبرا",
                "زايد 2",
                "حدائق الاهرام",
                "اكتوبر",
                "الشيخ زايد",
                "بالم هيلز",
                "سيتي ستارز",
                "هيليوبليس",
            ]
            base_columns = ["Barcode", "Product name", "SKU", "category", "PP"]
            used_branch_columns = set(alexandria_columns + ready_veg_columns) - set(base_columns)
            cairo_columns = base_columns + [
                col for col in pivot_df.columns if col not in used_branch_columns and col not in base_columns
            ]

            alexandria_df = pivot_df[[col for col in alexandria_columns if col in pivot_df.columns]]
            ready_veg_df = pivot_df[[col for col in ready_veg_columns if col in pivot_df.columns]]
            cairo_df = pivot_df[[col for col in cairo_columns if col in pivot_df.columns]]

            def reorder_columns(df):
                first_cols = ["Barcode", "SKU", "Product name"]
                last_cols = ["PP", "category"]
                middle_cols = sorted([col for col in df.columns if col not in first_cols + last_cols])
                ordered_cols = first_cols + middle_cols + last_cols
                return df[[col for col in ordered_cols if col in df.columns]]

            alexandria_df = reorder_columns(alexandria_df)
            ready_veg_df = reorder_columns(ready_veg_df)
            cairo_df = reorder_columns(cairo_df)

            category_order = {"فاكهه": 1, "خضار": 2, "جاهز": 3, "اعشاب": 4}

            def add_total_and_sort(df):
                fixed_cols = ["Barcode", "Product name", "SKU", "category", "PP"]
                branch_cols = [col for col in df.columns if col not in fixed_cols]
                df["total quantity"] = df[branch_cols].sum(axis=1)
                df["total"] = df["PP"] * df["total quantity"]
                df["category_order"] = df["category"].map(category_order)
                df = df.sort_values(by=["category_order", "Product name"])
                df = df.drop(columns=["category_order"])

                pp_index = df.columns.tolist().index("PP")
                cols = df.columns.tolist()
                cols.remove("total quantity")
                cols.remove("total")
                cols = cols[:pp_index] + ["total quantity", "PP", "total"] + cols[pp_index + 1 :]
                return df[cols]

            alexandria_df = add_total_and_sort(alexandria_df)
            ready_veg_df = add_total_and_sort(ready_veg_df)
            cairo_df = add_total_and_sort(cairo_df)

            def append_grand_total(df):
                # requires pandas imported as pd
                required = {"total quantity", "PP", "total"}
                if not required.issubset(df.columns):
                    return df

                cols = df.columns.tolist()
                try:
                    product_name_idx = cols.index("Product name")
                    pp_idx = cols.index("PP")
                except ValueError:
                    return df

                sum_columns = cols[product_name_idx + 1 : pp_idx]

                # compute number of columns with at least one non-null value (from the original df)
                num_nonempty_cols = int(df.notna().any(axis=0).sum())
                branches_count = max(0, num_nonempty_cols - 7)  # ensure non-negative integer

                # build grand total row
                grand_total_row = {col: "" for col in df.columns}
                grand_total_row["Product name"] = "Grand Total"

                for col in sum_columns:
                    if pd.api.types.is_numeric_dtype(df[col]):
                        grand_total_row[col] = df[col].sum()

                grand_total_row["total quantity"] = df["total quantity"].sum()
                grand_total_row["PP"] = df["PP"].sum()
                grand_total_row["total"] = df["total"].sum()

                df = pd.concat([df, pd.DataFrame([grand_total_row])], ignore_index=True)

                # build branch count row (عدد الفروع)
                branch_row = {col: "" for col in df.columns}
                branch_row["Product name"] = "عدد الفروع"

                # choose a sensible column to place the branches_count:
                target_col = None
                # prefer first numeric column in sum_columns
                for col in sum_columns:
                    if pd.api.types.is_numeric_dtype(df[col]):
                        target_col = col
                        break
                # fallback to "total quantity"
                if target_col is None and "total quantity" in df.columns and pd.api.types.is_numeric_dtype(df["total quantity"]):
                    target_col = "total quantity"
                # fallback to any numeric column
                if target_col is None:
                    for col in df.columns:
                        if pd.api.types.is_numeric_dtype(df[col]):
                            target_col = col
                            break
                # last resort: put into the first column after Product name (may be non-numeric)
                if target_col is None:
                    after_product_idx = product_name_idx + 1
                    target_col = cols[after_product_idx] if after_product_idx < len(cols) else df.columns[-1]

                branch_row[target_col] = int(branches_count)

                df = pd.concat([df, pd.DataFrame([branch_row])], ignore_index=True)
                return df


            alexandria_df = alexandria_df[alexandria_df["total"] != 0]
            ready_veg_df = ready_veg_df[ready_veg_df["total"] != 0]
            cairo_df = cairo_df[cairo_df["total"] != 0]

            alexandria_df = append_grand_total(alexandria_df)
            ready_veg_df = append_grand_total(ready_veg_df)
            cairo_df = append_grand_total(cairo_df)

            alex_buffer = BytesIO()
            with pd.ExcelWriter(alex_buffer, engine="xlsxwriter") as writer:
                alexandria_df.to_excel(writer, index=False)

            ready_buffer = BytesIO()
            with pd.ExcelWriter(ready_buffer, engine="xlsxwriter") as writer:
                ready_veg_df.to_excel(writer, index=False)

            cairo_buffer = BytesIO()
            with pd.ExcelWriter(cairo_buffer, engine="xlsxwriter") as writer:
                cairo_df.to_excel(writer, index=False)

        # Step 3: Assign invoice numbers to each branch-level XLSX
        special_branches = ["الابراهيميه", "سيدي بشر", "وينجت"]
        branch_offsets = {}
        filenames = [f for f in os.listdir(output_dir) if f.endswith(".xlsx")]
        file_branch_map = {filename: filename.split("_")[0] for filename in filenames}

        present_specials = [b for b in special_branches if b in file_branch_map.values()]
        other_branches = sorted(set(file_branch_map.values()) - set(special_branches))
        offset = 0
        for b in present_specials + other_branches:
            branch_offsets[b] = offset
            offset += 1

        for filename, branch_name in file_branch_map.items():
            final_invoice_number = base_invoice_number + branch_offsets.get(branch_name, 0)
            output_path = os.path.join(output_dir, filename)
            wb = load_workbook(output_path)
            if "فاتورة" in wb.sheetnames:
                ws = wb["فاتورة"]
                ws["E2"] = final_invoice_number
                wb.save(output_path)

        # Step 4: Consolidate all "فاتورة" sheets into one Workbook,
        # but only if at least one such sheet exists.
        invoice_filenames = []
        for filename in filenames:
            file_path = os.path.join(output_dir, filename)
            wb = load_workbook(file_path, data_only=True)
            if "فاتورة" in wb.sheetnames:
                invoice_filenames.append(filename)

        # Create consolidated workbook
        consolidated_wb = Workbook()
        if invoice_filenames:
            # Remove default empty sheet only if we'll add real sheets
            consolidated_wb.remove(consolidated_wb.active)
            for filename in invoice_filenames:
                file_path = os.path.join(output_dir, filename)
                wb = load_workbook(file_path, data_only=True)
                source_ws = wb["فاتورة"]
                new_sheet_name = os.path.splitext(filename)[0][:31]
                target_ws = consolidated_wb.create_sheet(title=new_sheet_name)

                # Copy merged cells & content + styles + dimensions
                for merged_range in source_ws.merged_cells.ranges:
                    target_ws.merge_cells(str(merged_range))
                for row in source_ws.iter_rows():
                    for cell in row:
                        new_cell = target_ws.cell(
                            row=cell.row, column=cell.column, value=cell.value
                        )
                        new_cell.font = Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            vertAlign=cell.font.vertAlign,
                            underline=cell.font.underline,
                            strike=cell.font.strike,
                            color=cell.font.color,
                        )
                        new_cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical, wrap_text=cell.alignment.wrap_text
                        )
                        new_cell.border = Border(
                            left=cell.border.left,
                            right=cell.border.right,
                            top=cell.border.top,
                            bottom=cell.border.bottom,
                        )
                        new_cell.fill = PatternFill(
                            fill_type=cell.fill.fill_type,
                            fgColor=cell.fill.fgColor,
                            bgColor=cell.fill.bgColor,
                        )
                        new_cell.number_format = cell.number_format

                # Copy row heights
                for row in source_ws.iter_rows():
                    target_ws.row_dimensions[row[0].row].height = source_ws.row_dimensions[row[0].row].height
                # Copy column widths
                for col in source_ws.columns:
                    col_letter = get_column_letter(col[0].column)
                    target_ws.column_dimensions[col_letter].width = source_ws.column_dimensions[col_letter].width
                # Add image if needed
                target_ws.add_image(Image("Picture1.png"), "A1")

        # If invoice_filenames was empty, consolidated_wb still has one default sheet—no removal needed.

        invoices_buffer = BytesIO()
        consolidated_wb.save(invoices_buffer)
        invoices_buffer.seek(0)

        # Step 5: Build PO summary & po_totals.xlsx
        po_summary = []
        excluded_files = {
            f"مجمع_طلبات_اسكندرية_{selected_date}.xlsx",
            f"مجمع_طلبات_الخضار_الجاهز_{selected_date}.xlsx",
            f"مجمع_طلبات_القاهرة_{selected_date}.xlsx",
            "فواتير.xlsx"
        }
        excluded_keywords = {"وينجت", "الابراهيميه", "سيدي بشر"}
        combined_wb = Workbook()
        combined_ws = combined_wb.active
        combined_ws.title = "CombinedOrders"
        current_row = 1
        first_file = True  # Add this before the loop that appends to combined_ws
        for filename in filenames:
            if filename in excluded_files or not filename.endswith(".xlsx"):
                continue
            file_path = os.path.join(output_dir, filename)
            wb = load_workbook(file_path, data_only=True)
            if "Sheet1" not in wb.sheetnames:
                continue
            ws = wb["Sheet1"]
            h1_text = ws["H1"].value
            total_col_idx = next(
                (
                    cell.column
                    for cell in ws[1]
                    if cell.value and str(cell.value).strip().lower() == "total"
                ),
                None,
            )
            if not total_col_idx:
                continue
            total_sum = 0
            for row in ws.iter_rows(min_row=2):
                val = row[total_col_idx - 1].value
                if isinstance(val, (int, float)):
                    total_sum += val

            invoice_number_val = None
            if "فاتورة" in wb.sheetnames:
                invoice_val = wb["فاتورة"]["E2"].value
                if isinstance(invoice_val, int):
                    invoice_number_val = invoice_val

            arabic_branch = filename.split("_")[0]
            english_branch = branches_translation_tlbt.get(arabic_branch, arabic_branch)
            po_summary.append((english_branch, arabic_branch, h1_text, total_sum, invoice_number_val))

        po_totals_wb = Workbook()
        po_ws = po_totals_wb.active
        po_ws.title = "Summary"
        po_ws.append(["branch (en)", "branch (ar)", "po", "Total of the po", "invoice_number"])
        for item in po_summary:
            po_ws.append(item)

        po_totals_buffer = BytesIO()
        po_totals_wb.save(po_totals_buffer)
        po_totals_buffer.seek(0)

        
        # Step 6: Build final ZIP containing:
        #   1) Inner ZIP of per-branch XLSX files (excluding excluded_files),
        #   2) po_totals,
        #   3) the three region-grouped xlsx buffers (alex/ready/cairo),
        #   4) "فواتير.xlsx" (the consolidated invoices workbook).
        output_zip_buffer = BytesIO()



        g1_insertions = []  # List to hold (G1_value, F_value)
        for filename in os.listdir(output_dir):
            # Skip non-xlsx or excluded by name or excluded by keyword
            if (
                not filename.endswith(".xlsx")
                or filename in excluded_files
                or any(kw in filename for kw in excluded_keywords)
            ):
                continue

            file_path = os.path.join(output_dir, filename)
            wb = load_workbook(file_path, data_only=True)

            # Skip if there’s no Sheet1
            if "Sheet1" not in wb.sheetnames:
                continue

            ws = wb["Sheet1"]
            h1_text = ws["H1"].value  # We’ll move this into F2 later

            # 4) Create a fresh in-memory Workbook to hold the “modified” version of Sheet1
            new_wb = Workbook()
            new_ws = new_wb.active

            # Copy all rows from original ws, but drop column 1 (“SKU”)
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx == 1:
                    # Header row: drop the very first cell
                    headers = list(row)[1:]
                    new_ws.append(headers)
                else:
                    # Data rows: again, skip the first column
                    new_ws.append(list(row)[1:])

            # 5) Find the indices of "Total", "Qty", and "Barcode" in the new_ws header
            header_cells = new_ws[1]  # tuple of Cell objects in row 1
            total_col_idx = None
            qty_col_idx = None
            barcode_col_idx = None

            for idx, cell in enumerate(header_cells, start=1):
                cell_val = cell.value
                if cell_val is None:
                    continue
                lower = str(cell_val).strip().lower()
                if lower == "total":
                    total_col_idx = idx
                elif lower == "qty":
                    qty_col_idx = idx
                elif lower == "barcode":
                    barcode_col_idx = idx

            # If there's no “Total” column, skip this file
            if total_col_idx is None:
                continue

            # 6) Sum up the “Total” column (numbers only), then append that sum as a new row
            total_sum = 0
            for row in new_ws.iter_rows(min_row=2, min_col=total_col_idx, max_col=total_col_idx):
                val = row[0].value
                if isinstance(val, (int, float)):
                    total_sum += val

            # Write the sum in the “Total” column of the first blank row
            first_blank_row = new_ws.max_row + 1
            total_col_letter = get_column_letter(total_col_idx)
            new_ws[f"{total_col_letter}{first_blank_row}"] = total_sum

            # In the row below the sum, put “*” under “Qty” if that column exists
            if qty_col_idx is not None:
                qty_col_letter = get_column_letter(qty_col_idx)
                new_ws[f"{qty_col_letter}{first_blank_row + 1}"] = "*"

            # 7) Move H1's value into cell F2
            new_ws["F2"] = h1_text

            # 8) Put the filename into G1 (one column to the right of “F”)
            new_ws["G1"] = filename.split("_")[0]
            g1_value = new_ws["G1"].value  # Store G1 for later

            # 9) Convert every “Barcode” cell (column) to a plain integer (no scientific notation)
            if barcode_col_idx is not None:
                for row in new_ws.iter_rows(min_row=2, min_col=barcode_col_idx, max_col=barcode_col_idx):
                    cell = row[0]
                    cell_val = cell.value
                    if isinstance(cell_val, (int, float, str)):
                        try:
                            int_val = int(float(cell_val))
                            cell.value = int_val
                            cell.number_format = '0'  # No decimal places
                        except:
                            pass
            
            
            # 10) Append ALL of new_ws’s rows into combined_ws
            if first_file:
                combined_ws.append([])  # Adds an empty row before the first table
                first_file = False
                
            for row_idx, row in enumerate(new_ws.iter_rows(values_only=True), start=1):
                if row_idx == 1:
                    continue  # Skip header
                combined_ws.append(row)
                # Check F column (column index 6, since you dropped 1st column)
                if len(row) >= 6 and row[5]:  # 0-based index: column F is index 5
                    current_combined_row = combined_ws.max_row
                    g1_insertions.append((g1_value, current_combined_row - 1))  # Place G1 in column G (index 6), row above
        for g1_value, f_row in g1_insertions:
            combined_ws[f"G{f_row}"] = g1_value

        final_combined_buffer = BytesIO()
        combined_wb.save(final_combined_buffer)
        final_combined_buffer.seek(0)


        output_zip_buffer = BytesIO()

        excluded_files = {
                    f"po_totals_{selected_date}.xlsx",
                    f"مجمع_طلبات_اسكندرية_{selected_date}.xlsx",
                    f"مجمع_طلبات_الخضار_الجاهز_{selected_date}.xlsx",
                    f"مجمع_طلبات_القاهرة_{selected_date}.xlsx",
                    "فواتير.xlsx",
                    f"طلبيات_{selected_date}.xlsx"
                }


        with zipfile.ZipFile(output_zip_buffer, "w") as zipf:
            inner_zip_buffer = BytesIO()
            with zipfile.ZipFile(inner_zip_buffer, "w") as inner_zip:
                for excel_file in os.listdir(output_dir):
                    if excel_file not in excluded_files and excel_file.endswith(".xlsx"):
                        excel_path = os.path.join(output_dir, excel_file)
                        inner_zip.write(excel_path, arcname=excel_file)
            inner_zip_buffer.seek(0)
            zipf.writestr(f"ملفات الفروع_{selected_date}.zip", inner_zip_buffer.getvalue())
            zipf.writestr(f"po_totals_{selected_date}.xlsx", po_totals_buffer.getvalue())
            zipf.writestr(f"مجمع_طلبات_اسكندرية_{selected_date}.xlsx", alex_buffer.getvalue())
            zipf.writestr(f"مجمع_طلبات_الخضار_الجاهز_{selected_date}.xlsx", ready_buffer.getvalue())
            zipf.writestr(f"مجمع_طلبات_القاهرة_{selected_date}.xlsx", cairo_buffer.getvalue())
            zipf.writestr("فواتير.xlsx", invoices_buffer.getvalue())
            zipf.writestr(f"طلبيات_{selected_date}.xlsx", final_combined_buffer.getvalue())

        output_zip_buffer.seek(0)
        return output_zip_buffer.getvalue(), offset
